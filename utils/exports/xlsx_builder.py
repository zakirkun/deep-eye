"""Excel xlsx workbook builder for executive distribution.

Lazy-loads openpyxl. Five sheets: Summary, Vulnerabilities, Reconnaissance, CVEs, Compliance (placeholder).
"""
import sys
import subprocess
from datetime import datetime, timezone
from typing import Dict, List


SEVERITY_COLORS = {
    "critical": ("8B0000", "FFFFFF"),  # bg, font
    "high": ("FF4500", "FFFFFF"),
    "medium": ("FFA500", "000000"),
    "low": ("FFD700", "000000"),
    "info": ("87CEEB", "000000"),
}

VULN_COLUMNS = [
    "type",
    "severity",
    "url",
    "parameter",
    "payload",
    "evidence",
    "description",
    "remediation",
    "cve_references",
    "cvss_score",
    "timestamp",
]

COMPLIANCE_HEADERS = [
    "framework",
    "control_id",
    "vulnerability_type",
    "severity",
    "status",
]

CVE_HEADERS = ["cve_id", "vulnerability_type", "url", "severity", "cvss_score"]

_EVIDENCE_LIMIT = 1000


def _ensure_openpyxl(interactive: bool = True) -> bool:
    """Try to import openpyxl. Prompt for install if missing."""
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        pass

    if not interactive or not sys.stdin.isatty():
        return False

    try:
        answer = input("[!] xlsx export needs openpyxl. Install now? [y/N]: ")
    except (EOFError, KeyboardInterrupt):
        return False

    if answer.strip().lower() not in ("y", "yes"):
        return False

    try:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            check=True,
        )
        import openpyxl  # noqa: F401
        return True
    except (subprocess.CalledProcessError, ImportError):
        return False


def _flatten_cve(refs) -> str:
    if not refs:
        return ""
    if isinstance(refs, list):
        return "; ".join(str(r) for r in refs)
    return str(refs)


def _build_summary_sheet(ws, results: Dict) -> None:
    """Populate Summary sheet."""
    from openpyxl.styles import Font

    severity_counts = results.get("severity_summary", {})
    rows = [
        ("Field", "Value"),
        ("Target", str(results.get("target", "N/A"))),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Scan Duration", str(results.get("duration", "N/A"))),
        ("URLs Crawled", results.get("urls_crawled", 0)),
        ("Total Vulnerabilities", len(results.get("vulnerabilities", []))),
        ("Critical", severity_counts.get("critical", 0)),
        ("High", severity_counts.get("high", 0)),
        ("Medium", severity_counts.get("medium", 0)),
        ("Low", severity_counts.get("low", 0)),
        ("Info", severity_counts.get("info", 0)),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _build_vulns_sheet(ws, vulnerabilities: List[Dict], timestamp: str) -> None:
    """Populate Vulnerabilities sheet with severity-colored rows."""
    from openpyxl.styles import Font, PatternFill

    ws.append(VULN_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for vuln in vulnerabilities:
        evidence = str(vuln.get("evidence", ""))[:_EVIDENCE_LIMIT]
        row = [
            str(vuln.get("type", "")),
            str(vuln.get("severity", "")),
            str(vuln.get("url", "")),
            str(vuln.get("parameter", "")),
            str(vuln.get("payload", "")),
            evidence,
            str(vuln.get("description", "")),
            str(vuln.get("remediation", "")),
            _flatten_cve(vuln.get("cve_references", [])),
            str(vuln.get("cvss_score", "")),
            timestamp,
        ]
        ws.append(row)

        severity = str(vuln.get("severity", "info")).lower()
        bg, fg = SEVERITY_COLORS.get(severity, SEVERITY_COLORS["info"])
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        font = Font(color=fg)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = font

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _build_recon_sheet(ws, recon: Dict) -> None:
    """Populate Reconnaissance sheet."""
    from openpyxl.styles import Font

    if not recon:
        ws.append(["No reconnaissance data"])
        return

    ws.append(["Category", "Key", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    dns = recon.get("dns", {})
    if dns:
        for record_type, records in dns.items():
            if records:
                ws.append(
                    [
                        "DNS",
                        str(record_type).upper(),
                        ", ".join(str(r) for r in records),
                    ]
                )
        ws.append([])  # blank row separator

    osint = recon.get("osint", {})
    if osint:
        for key in ("emails", "subdomains", "github_leaks", "breaches"):
            values = osint.get(key, [])
            if values:
                ws.append(
                    ["OSINT", key, ", ".join(str(v) for v in values[:50])]
                )
        ws.append([])

    techs = recon.get("technologies", [])
    if techs:
        ws.append(
            ["Technologies", "detected", ", ".join(str(t) for t in techs)]
        )


def _build_cves_sheet(ws, vulnerabilities: List[Dict]) -> None:
    """Populate CVEs sheet — pivot of cve_references field."""
    from openpyxl.styles import Font

    ws.append(CVE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for vuln in vulnerabilities:
        refs = vuln.get("cve_references", [])
        if not refs:
            continue
        if not isinstance(refs, list):
            refs = [refs]
        for cve in refs:
            ws.append(
                [
                    str(cve),
                    str(vuln.get("type", "")),
                    str(vuln.get("url", "")),
                    str(vuln.get("severity", "")),
                    str(vuln.get("cvss_score", "")),
                ]
            )


def _build_compliance_sheet(ws) -> None:
    """Populate Compliance sheet — header only, body filled by Group B."""
    from openpyxl.styles import Font

    ws.append(COMPLIANCE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_xlsx(results: Dict, output_path: str, interactive: bool = True) -> bool:
    """Build xlsx workbook from scan results.

    Args:
        results: Scan result dict.
        output_path: Path to write .xlsx file.
        interactive: If True, prompt to install openpyxl when missing.

    Returns:
        True on success, False if openpyxl unavailable or write failed.
    """
    if not _ensure_openpyxl(interactive=interactive):
        return False

    from openpyxl import Workbook

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    vulnerabilities = results.get("vulnerabilities", [])
    recon = results.get("reconnaissance", {})

    summary_ws = wb.create_sheet("Summary")
    _build_summary_sheet(summary_ws, results)

    vulns_ws = wb.create_sheet("Vulnerabilities")
    _build_vulns_sheet(vulns_ws, vulnerabilities, timestamp)

    recon_ws = wb.create_sheet("Reconnaissance")
    _build_recon_sheet(recon_ws, recon)

    cves_ws = wb.create_sheet("CVEs")
    _build_cves_sheet(cves_ws, vulnerabilities)

    compliance_ws = wb.create_sheet("Compliance")
    _build_compliance_sheet(compliance_ws)

    try:
        wb.save(output_path)
        return True
    except Exception:
        return False
