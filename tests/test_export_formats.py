"""Tests for export format builders (JUnit XML, CSV, Excel)."""
import csv
import io
import os
import sys
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from unittest.mock import patch

import pytest

# Make repo root importable
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from utils.exports.junit_builder import build_junit_xml
from utils.exports.csv_builder import build_csv, COLUMNS as CSV_COLUMNS


SAMPLE_VULN = {
    "type": "SQL Injection",
    "severity": "critical",
    "url": "https://example.com/login",
    "parameter": "user",
    "payload": "' OR 1=1--",
    "evidence": "MySQL syntax error in response",
    "description": "SQLi in login form",
    "remediation": "Use parameterized queries",
    "cve_references": ["CVE-2021-XXXX"],
    "cvss_score": 9.8,
}

SAMPLE_VULN_XSS = {
    "type": "XSS",
    "severity": "high",
    "url": "https://example.com/search",
    "parameter": "q",
    "payload": "<script>alert(1)</script>",
    "evidence": "Reflected, contains <script> tags & special chars",
    "description": "Reflected XSS",
    "remediation": "Encode output",
    "cve_references": [],
    "cvss_score": 7.5,
}

SAMPLE_RESULTS = {
    "target": "https://example.com",
    "duration": 12.34,
    "vulnerabilities": [SAMPLE_VULN, SAMPLE_VULN_XSS],
    "severity_summary": {"critical": 1, "high": 1, "medium": 0, "low": 0, "info": 0},
    "urls_crawled": 42,
    "reconnaissance": {
        "dns": {"a": ["1.2.3.4"], "mx": ["mail.example.com"]},
        "osint": {
            "emails": ["admin@example.com"],
            "subdomains": ["api.example.com"],
        },
        "technologies": ["nginx", "PHP"],
    },
}

EMPTY_RESULTS = {
    "target": "https://example.com",
    "duration": 1.0,
    "vulnerabilities": [],
    "severity_summary": {},
    "urls_crawled": 0,
}


# ---------- JUnit XML ----------

class TestJUnitXML:
    def test_basic(self):
        xml_bytes = build_junit_xml(SAMPLE_RESULTS)
        root = ET.fromstring(xml_bytes)
        assert root.tag == "testsuites"
        assert root.attrib["tests"] == "2"
        assert root.attrib["failures"] == "2"

        suite = root.find("testsuite")
        assert suite is not None
        assert suite.attrib["name"] == "https://example.com"

        cases = suite.findall("testcase")
        assert len(cases) == 2
        assert cases[0].attrib["classname"] == "SQL Injection"
        assert cases[0].attrib["name"] == "https://example.com/login [user]"

        failure = cases[0].find("failure")
        assert failure is not None
        assert failure.attrib["type"] == "critical"
        assert "SQL Injection" in failure.attrib["message"]
        assert "Payload:" in failure.text
        assert "Evidence:" in failure.text
        assert "CVE-2021-XXXX" in failure.text

    def test_empty(self):
        xml_bytes = build_junit_xml(EMPTY_RESULTS)
        root = ET.fromstring(xml_bytes)
        assert root.attrib["tests"] == "0"
        assert root.attrib["failures"] == "0"
        suite = root.find("testsuite")
        assert suite is not None
        assert len(suite.findall("testcase")) == 0

    def test_xml_escaping(self):
        xml_bytes = build_junit_xml({"target": "x", "duration": 0, "vulnerabilities": [SAMPLE_VULN_XSS]})
        root = ET.fromstring(xml_bytes)
        failure = root.find(".//failure")
        assert failure is not None
        assert "<script>alert(1)</script>" in failure.text

    def test_no_parameter(self):
        vuln = dict(SAMPLE_VULN)
        vuln["parameter"] = ""
        xml_bytes = build_junit_xml({"target": "x", "duration": 0, "vulnerabilities": [vuln]})
        root = ET.fromstring(xml_bytes)
        case = root.find(".//testcase")
        assert case.attrib["name"] == "https://example.com/login"


# ---------- CSV ----------

class TestCSV:
    def test_columns(self):
        text = build_csv(SAMPLE_RESULTS)
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        header = next(reader)
        assert header == CSV_COLUMNS
        rows = list(reader)
        assert len(rows) == 2

    def test_quoting(self):
        text = build_csv({"vulnerabilities": [SAMPLE_VULN_XSS]})
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        next(reader)
        row = next(reader)
        assert "Reflected, contains <script> tags & special chars" in row[5]

    def test_utf8_bom(self):
        text = build_csv(SAMPLE_RESULTS)
        encoded = text.encode("utf-8")
        assert encoded.startswith(b"\xef\xbb\xbf")

    def test_empty(self):
        text = build_csv(EMPTY_RESULTS)
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        header = next(reader)
        assert header == CSV_COLUMNS
        assert list(reader) == []

    def test_cve_list_flattened(self):
        text = build_csv({"vulnerabilities": [SAMPLE_VULN]})
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        next(reader)
        row = next(reader)
        assert row[8] == "CVE-2021-XXXX"


# ---------- xlsx ----------

class TestXLSX:
    def _build(self, results=None):
        from utils.exports.xlsx_builder import build_xlsx
        results = results or SAMPLE_RESULTS
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            path = tf.name
        success = build_xlsx(results, path, interactive=False)
        return success, path

    def test_skip_when_openpyxl_missing(self):
        from utils.exports.xlsx_builder import _ensure_openpyxl
        # Patch import to fail for openpyxl
        import builtins
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("mocked missing")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=fake_import):
            assert _ensure_openpyxl(interactive=False) is False

    def test_basic_workbook(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        success, path = self._build()
        assert success
        assert os.path.exists(path)

        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert wb.sheetnames == ["Summary", "Vulnerabilities", "Reconnaissance", "CVEs", "Compliance"]
        os.unlink(path)

    def test_compliance_placeholder(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        success, path = self._build()
        from openpyxl import load_workbook
        wb = load_workbook(path)
        compliance = wb["Compliance"]
        assert compliance.max_row == 1
        headers = [c.value for c in compliance[1]]
        assert headers == ["framework", "control_id", "vulnerability_type", "severity", "status"]
        os.unlink(path)

    def test_severity_colors(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        success, path = self._build()
        from openpyxl import load_workbook
        wb = load_workbook(path)
        vulns = wb["Vulnerabilities"]
        cell = vulns.cell(row=2, column=1)
        rgb = cell.fill.start_color.rgb or ""
        assert "8B0000" in str(rgb)
        os.unlink(path)

    def test_cves_sheet(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        success, path = self._build()
        from openpyxl import load_workbook
        wb = load_workbook(path)
        cves = wb["CVEs"]
        assert cves.max_row == 2
        assert cves.cell(row=2, column=1).value == "CVE-2021-XXXX"
        os.unlink(path)
